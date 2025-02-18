from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = os.path.expanduser("~/secure_data/responses.xlsx")
OWNER_SECRET = "your_strong_secret_key"  # Change this to a secure key

# Ensure the Excel file exists with headers
if not os.path.exists(EXCEL_FILE):
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(["Game Type", "Email", "Timer Value"])  # Updated Header row
    wb.save(EXCEL_FILE)

@app.route('/store-data', methods=['POST'])
def store_data():
    data = request.json
    game_type = data.get('game_type')  # "sudoku" or "crossword"
    email = data.get('email')
    timervalue = data.get('timervalue')

    if not game_type or game_type not in ["sudoku", "crossword"]:
        return jsonify({'message': 'Game Type must be "sudoku" or "crossword"!'}), 400
    if not email or not timervalue:
        return jsonify({'message': 'Email and Timer Value are required!'}), 400

    # Load the Excel file
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Check if (game_type, email) pair already exists
    existing_entries = {(row[0], row[1]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0] and row[1]}

    if (game_type, email) in existing_entries:
        return jsonify({'message': 'This game and email combination already exists. Data not stored.'}), 409  # Conflict

    # Store new entry
    ws.append([game_type, email, timervalue])
    wb.save(EXCEL_FILE)

    return jsonify({'message': 'Data stored successfully!'})

@app.route('/get-data', methods=['GET'])
def get_data():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    data = [{"game_type": row[0], "email": row[1], "timervalue": row[2]} for row in ws.iter_rows(min_row=2, values_only=True)]

    return jsonify(data)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
